from __future__ import annotations

import argparse
import hashlib
import json
import math
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries

SUPPORTED_TYPES = {"string", "integer", "number", "local_datetime"}
FORMULA_POLICIES = {"forbid_in_range", "allow_with_cached_value"}
LAYOUT_POLICIES = {"forbid_in_range", "report_only"}


class ExcelAuditError(ValueError):
    """Raised when the workbook or extraction specification cannot be audited."""


def parse_range(value: Any) -> tuple[int, int, int, int]:
    if not isinstance(value, str) or not value:
        raise ExcelAuditError("boundary.range must be a non-empty A1 range")
    try:
        bounds = range_boundaries(value)
    except (TypeError, ValueError) as error:
        raise ExcelAuditError(f"invalid boundary.range: {value!r}") from error
    if bounds[1] == bounds[3]:
        raise ExcelAuditError("boundary.range must contain a header and at least one data row")
    return bounds


def load_spec(path: str | Path) -> dict[str, Any]:
    spec_path = Path(path)
    if not spec_path.is_file():
        raise ExcelAuditError(f"spec file does not exist: {spec_path}")
    try:
        spec = json.loads(spec_path.read_text(encoding="utf-8"))
    except (OSError, UnicodeError) as error:
        raise ExcelAuditError(f"cannot read spec: {error}") from error
    except json.JSONDecodeError as error:
        raise ExcelAuditError(f"invalid spec JSON: {error.msg}") from error
    if not isinstance(spec, dict):
        raise ExcelAuditError("spec root must be an object")

    required = {
        "sheet",
        "boundary",
        "columns",
        "key",
        "formula_policy",
        "layout_policy",
    }
    missing = required - set(spec)
    if missing:
        raise ExcelAuditError(f"spec misses fields: {sorted(missing)}")
    if not isinstance(spec["sheet"], str) or not spec["sheet"]:
        raise ExcelAuditError("sheet must be a non-empty string")

    boundary = spec["boundary"]
    if not isinstance(boundary, dict) or boundary.get("mode") != "fixed_range":
        raise ExcelAuditError("boundary.mode must be fixed_range")
    if not isinstance(boundary.get("reject_nonempty_below"), bool):
        raise ExcelAuditError("boundary.reject_nonempty_below must be boolean")
    min_col, _, max_col, _ = parse_range(boundary.get("range"))

    columns = spec["columns"]
    if not isinstance(columns, dict) or not columns:
        raise ExcelAuditError("columns must be a non-empty object")
    if max_col - min_col + 1 != len(columns):
        raise ExcelAuditError("range width must match the number of declared columns")
    for name, rule in columns.items():
        if not isinstance(name, str) or not name:
            raise ExcelAuditError("column names must be non-empty strings")
        if not isinstance(rule, dict):
            raise ExcelAuditError(f"column {name} must be an object")
        if rule.get("type") not in SUPPORTED_TYPES:
            raise ExcelAuditError(f"column {name} has unsupported type: {rule.get('type')!r}")
        if not isinstance(rule.get("nullable"), bool):
            raise ExcelAuditError(f"column {name} must declare nullable as boolean")
        if rule["type"] == "local_datetime":
            timezone = rule.get("source_timezone")
            if not isinstance(timezone, str) or not timezone:
                raise ExcelAuditError(f"local_datetime column {name} must declare source_timezone")
            try:
                ZoneInfo(timezone)
            except ZoneInfoNotFoundError as error:
                raise ExcelAuditError(
                    f"local_datetime column {name} has unknown source_timezone: {timezone}"
                ) from error

    key = spec["key"]
    if not isinstance(key, list) or not key or len(key) != len(set(key)):
        raise ExcelAuditError("key must be a non-empty list of unique column names")
    if not all(isinstance(name, str) and name in columns for name in key):
        raise ExcelAuditError("every key column must exist in columns")
    if spec["formula_policy"] not in FORMULA_POLICIES:
        raise ExcelAuditError("formula_policy must be forbid_in_range or allow_with_cached_value")
    if spec["layout_policy"] not in LAYOUT_POLICIES:
        raise ExcelAuditError("layout_policy must be forbid_in_range or report_only")
    return spec


def json_value(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def is_blank(value: Any) -> bool:
    return value is None or value == ""


def selected_values(sheet: Any, cell_range: str) -> list[list[Any]]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    return [
        [sheet.cell(row=row, column=column).value for column in range(min_col, max_col + 1)]
        for row in range(min_row, max_row + 1)
    ]


def value_matches_type(value: Any, declared_type: str) -> bool:
    if declared_type == "string":
        return isinstance(value, str)
    if declared_type == "integer":
        return isinstance(value, int) and not isinstance(value, bool)
    if declared_type == "number":
        return (
            isinstance(value, (int, float, Decimal))
            and not isinstance(value, bool)
            and math.isfinite(float(value))
        )
    if declared_type == "local_datetime":
        return isinstance(value, datetime) and value.tzinfo is None
    return False


def inspect_schema(
    rows: list[list[Any]],
    *,
    first_excel_row: int,
    columns: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    report: dict[str, Any] = {}
    for index, (name, rule) in enumerate(columns.items()):
        null_rows = []
        invalid_rows = []
        normalized_preview = []
        for offset, row in enumerate(rows):
            excel_row = first_excel_row + offset
            value = row[index]
            if is_blank(value):
                null_rows.append(excel_row)
                continue
            if not value_matches_type(value, rule["type"]):
                invalid_rows.append(excel_row)
                continue
            if rule["type"] == "local_datetime" and len(normalized_preview) < 2:
                localized = value.replace(tzinfo=ZoneInfo(rule["source_timezone"]))
                normalized_preview.append(localized.isoformat())
        valid = not invalid_rows and (rule["nullable"] or not null_rows)
        report[name] = {
            "type": rule["type"],
            "nullable": rule["nullable"],
            "null_rows": null_rows,
            "invalid_type_rows": invalid_rows,
            "valid": valid,
        }
        if "source_timezone" in rule:
            report[name]["source_timezone"] = rule["source_timezone"]
            report[name]["localized_preview"] = normalized_preview
    return report


def inspect_key(
    rows: list[list[Any]],
    *,
    first_excel_row: int,
    column_names: list[str],
    key: list[str],
) -> dict[str, Any]:
    indexes = [column_names.index(name) for name in key]
    first_seen: dict[tuple[Any, ...], int] = {}
    null_rows = []
    duplicate_rows = []
    for offset, row in enumerate(rows):
        excel_row = first_excel_row + offset
        value = tuple(row[index] for index in indexes)
        if any(is_blank(part) for part in value):
            null_rows.append(excel_row)
        elif value in first_seen:
            duplicate_rows.append(
                {"row": excel_row, "duplicates_row": first_seen[value], "value": list(value)}
            )
        else:
            first_seen[value] = excel_row
    return {
        "columns": key,
        "null_rows": null_rows,
        "duplicate_rows": duplicate_rows,
        "valid": not null_rows and not duplicate_rows,
    }


def ranges_overlap(
    first: tuple[int, int, int, int],
    second: tuple[int, int, int, int],
) -> bool:
    first_min_col, first_min_row, first_max_col, first_max_row = first
    second_min_col, second_min_row, second_max_col, second_max_row = second
    return not (
        first_max_col < second_min_col
        or second_max_col < first_min_col
        or first_max_row < second_min_row
        or second_max_row < first_min_row
    )


def inspect_with_pandas(path: Path, spec: dict[str, Any]) -> dict[str, Any]:
    cell_range = spec["boundary"]["range"]
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    usecols = f"{get_column_letter(min_col)}:{get_column_letter(max_col)}"
    try:
        frame = pd.read_excel(
            path,
            sheet_name=spec["sheet"],
            header=min_row - 1,
            usecols=usecols,
            nrows=max_row - min_row,
            dtype=object,
            engine="openpyxl",
            keep_default_na=False,
        )
    except Exception as error:
        return {"loaded": False, "error": str(error), "valid": False}
    columns = [str(column) for column in frame.columns]
    expected_columns = list(spec["columns"])
    expected_rows = max_row - min_row
    return {
        "loaded": True,
        "rows": len(frame),
        "columns": columns,
        "dtypes": {str(name): str(dtype) for name, dtype in frame.dtypes.items()},
        "preview": [
            {str(key): json_value(value) for key, value in row.items()}
            for row in frame.head(2).to_dict(orient="records")
        ],
        "valid": columns == expected_columns and len(frame) == expected_rows,
    }


def open_workbook(path: Path, *, data_only: bool) -> Any:
    try:
        return load_workbook(path, read_only=False, data_only=data_only)
    except Exception as error:
        raise ExcelAuditError(f"cannot open workbook: {error}") from error


def audit_workbook(input_path: str | Path, spec_path: str | Path) -> dict[str, Any]:
    path = Path(input_path)
    if not path.is_file():
        raise ExcelAuditError(f"workbook does not exist: {path}")
    spec = load_spec(spec_path)
    workbook = open_workbook(path, data_only=False)
    cached_workbook = open_workbook(path, data_only=True)
    sheet_names = workbook.sheetnames
    if spec["sheet"] not in sheet_names:
        failed = ["sheet_exists"]
        return {
            "file": {"path": str(path), "sha256": hashlib.sha256(path.read_bytes()).hexdigest()},
            "workbook": {"sheet_names": sheet_names},
            "selection": {"valid": False, "error": f"missing sheet: {spec['sheet']}"},
            "pandas": {"loaded": False, "valid": False},
            "summary": {
                "valid": False,
                "failed_checks": failed,
                "failed_check_count": len(failed),
            },
        }

    sheet = workbook[spec["sheet"]]
    cached_sheet = cached_workbook[spec["sheet"]]
    cell_range = spec["boundary"]["range"]
    bounds = range_boundaries(cell_range)
    min_col, min_row, max_col, max_row = bounds
    values_with_formulas = selected_values(sheet, cell_range)
    values_with_cache = selected_values(cached_sheet, cell_range)
    header = values_with_formulas[0]
    expected_header = list(spec["columns"])
    header_matches = header == expected_header

    formulas_in_range = []
    cached_values: dict[str, Any] = {}
    missing_cached_values = []
    for row in range(min_row, max_row + 1):
        for column in range(min_col, max_col + 1):
            cell = sheet.cell(row=row, column=column)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas_in_range.append(cell.coordinate)
                cached = cached_sheet.cell(row=row, column=column).value
                cached_values[cell.coordinate] = json_value(cached)
                if cached is None:
                    missing_cached_values.append(cell.coordinate)

    if spec["formula_policy"] == "forbid_in_range":
        formula_valid = not formulas_in_range
        data_rows = values_with_formulas[1:]
    else:
        formula_valid = not missing_cached_values
        data_rows = values_with_cache[1:]

    spill_rows = []
    if spec["boundary"]["reject_nonempty_below"]:
        for row in range(max_row + 1, sheet.max_row + 1):
            cells = [sheet.cell(row=row, column=column) for column in range(min_col, max_col + 1)]
            if any(not is_blank(cell.value) for cell in cells):
                spill_rows.append(
                    {
                        "row": row,
                        "nonempty_cells": [
                            cell.coordinate for cell in cells if not is_blank(cell.value)
                        ],
                    }
                )
    boundary_valid = not spill_rows

    hidden_rows = [row for row in range(min_row, max_row + 1) if sheet.row_dimensions[row].hidden]
    hidden_columns = [
        get_column_letter(column)
        for column in range(min_col, max_col + 1)
        if sheet.column_dimensions[get_column_letter(column)].hidden
    ]
    merged_intersections = [
        str(merged) for merged in sheet.merged_cells.ranges if ranges_overlap(bounds, merged.bounds)
    ]
    layout_valid = spec["layout_policy"] == "report_only" or not (
        hidden_rows or hidden_columns or merged_intersections
    )

    schema = inspect_schema(
        data_rows,
        first_excel_row=min_row + 1,
        columns=spec["columns"],
    )
    schema_valid = header_matches and all(column["valid"] for column in schema.values())
    key = inspect_key(
        data_rows,
        first_excel_row=min_row + 1,
        column_names=expected_header,
        key=spec["key"],
    )
    pandas_report = inspect_with_pandas(path, spec)

    checks = {
        "header_matches": header_matches,
        "boundary_valid": boundary_valid,
        "formula_policy_valid": formula_valid,
        "layout_policy_valid": layout_valid,
        "schema_valid": schema_valid,
        "key_valid": key["valid"],
        "pandas_selection_valid": pandas_report["valid"],
    }
    failed = [name for name, valid in checks.items() if not valid]
    formulas_all = [
        cell.coordinate
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    all_hidden_rows = [row for row, dimension in sheet.row_dimensions.items() if dimension.hidden]
    all_hidden_columns = [
        name for name, dimension in sheet.column_dimensions.items() if dimension.hidden
    ]
    return {
        "file": {
            "path": str(path),
            "bytes": path.stat().st_size,
            "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
        },
        "workbook": {
            "sheet_names": sheet_names,
            "active_sheet": workbook.active.title,
            "merged_ranges": [str(value) for value in sheet.merged_cells.ranges],
            "hidden_rows": all_hidden_rows,
            "hidden_columns": all_hidden_columns,
            "formulas": formulas_all,
        },
        "selection": {
            "sheet": spec["sheet"],
            "range": cell_range,
            "header": header,
            "expected_header": expected_header,
            "header_matches": header_matches,
            "data_rows": len(data_rows),
            "spill_rows": spill_rows,
            "hidden_rows": hidden_rows,
            "hidden_columns": hidden_columns,
            "merged_intersections": merged_intersections,
            "valid": header_matches and boundary_valid and layout_valid,
        },
        "formulas": {
            "policy": spec["formula_policy"],
            "in_range": formulas_in_range,
            "cached_values": cached_values,
            "missing_cached_values": missing_cached_values,
            "valid": formula_valid,
        },
        "schema": {
            "columns": schema,
            "key": key,
            "valid": schema_valid and key["valid"],
        },
        "pandas": pandas_report,
        "checks": checks,
        "summary": {
            "valid": not failed,
            "failed_checks": failed,
            "failed_check_count": len(failed),
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Audit an Excel extraction range")
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--spec", required=True, type=Path)
    parser.add_argument("--allow-failures", action="store_true")
    args = parser.parse_args()
    try:
        report = audit_workbook(args.input, args.spec)
    except ExcelAuditError as error:
        print(json.dumps({"error": str(error)}, ensure_ascii=False))
        raise SystemExit(2) from error
    json.dump(report, sys.stdout, ensure_ascii=False, indent=2)
    sys.stdout.write("\n")
    if not report["summary"]["valid"] and not args.allow_failures:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
