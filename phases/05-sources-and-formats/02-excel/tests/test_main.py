from __future__ import annotations

import importlib.util
import json
import subprocess
import sys
import unittest
from collections.abc import Callable
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
ARTIFACT = ROOT / "outputs" / "excel_audit.py"
DATA = ROOT.parent / "data"
SPEC_PATH = DATA / "excel_spec.json"
VALID = DATA / "tiny" / "orders_report.xlsx"
SHIFTED = DATA / "tiny" / "orders_report_shifted.xlsx"
MODULE_SPEC = importlib.util.spec_from_file_location("excel_audit", ARTIFACT)
if MODULE_SPEC is None or MODULE_SPEC.loader is None:
    raise RuntimeError(f"cannot load {ARTIFACT}")
AUDITOR = importlib.util.module_from_spec(MODULE_SPEC)
MODULE_SPEC.loader.exec_module(AUDITOR)


class ExcelAuditTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.report = AUDITOR.audit_workbook(VALID, SPEC_PATH)

    def audit_modified(
        self,
        mutate_workbook: Callable[[Any], None],
        *,
        mutate_spec: Callable[[dict[str, Any]], None] | None = None,
    ) -> dict[str, Any]:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            workbook_path = root / "orders.xlsx"
            spec_path = root / "spec.json"
            workbook = load_workbook(VALID, data_only=False)
            mutate_workbook(workbook)
            workbook.save(workbook_path)
            spec = json.loads(SPEC_PATH.read_text(encoding="utf-8"))
            if mutate_spec is not None:
                mutate_spec(spec)
            spec_path.write_text(json.dumps(spec), encoding="utf-8")
            return AUDITOR.audit_workbook(workbook_path, spec_path)

    def test_declared_range_is_valid(self) -> None:
        self.assertTrue(self.report["summary"]["valid"])
        self.assertEqual(self.report["selection"]["data_rows"], 5)
        self.assertEqual(self.report["selection"]["header"][0], "order_id")
        self.assertEqual(self.report["summary"]["failed_checks"], [])

    def test_workbook_structure_is_visible_but_outside_range(self) -> None:
        self.assertEqual(
            self.report["workbook"]["sheet_names"],
            ["Инструкция", "Заказы", "Справочник"],
        )
        self.assertIn("A1:G1", self.report["workbook"]["merged_ranges"])
        self.assertEqual(self.report["workbook"]["hidden_columns"], ["H"])
        self.assertEqual(self.report["selection"]["merged_intersections"], [])
        self.assertEqual(self.report["selection"]["hidden_columns"], [])

    def test_formula_text_is_visible_but_excluded_from_range(self) -> None:
        self.assertIn("G5", self.report["workbook"]["formulas"])
        self.assertEqual(self.report["formulas"]["in_range"], [])
        self.assertTrue(self.report["formulas"]["valid"])

    def test_shifted_header_fails_explicitly(self) -> None:
        report = AUDITOR.audit_workbook(SHIFTED, SPEC_PATH)
        self.assertFalse(report["summary"]["valid"])
        self.assertFalse(report["selection"]["header_matches"])
        self.assertIn("header_matches", report["summary"]["failed_checks"])
        self.assertEqual(
            report["selection"]["header"][0],
            "Сформировано повторно: структура сдвинута",
        )

    def test_extra_row_below_fixed_range_is_not_silently_truncated(self) -> None:
        def add_order(workbook: Any) -> None:
            sheet = workbook["Заказы"]
            for column, value in enumerate(
                ["O2199", "U099", sheet["C9"].value, 10.0, "RUB", 1],
                start=1,
            ):
                sheet.cell(row=10, column=column, value=value)

        report = self.audit_modified(add_order)
        self.assertFalse(report["summary"]["valid"])
        self.assertEqual(report["selection"]["spill_rows"][0]["row"], 10)
        self.assertIn("boundary_valid", report["summary"]["failed_checks"])

    def test_hidden_row_inside_range_is_rejected(self) -> None:
        def hide_row(workbook: Any) -> None:
            workbook["Заказы"].row_dimensions[6].hidden = True

        report = self.audit_modified(hide_row)
        self.assertFalse(report["summary"]["valid"])
        self.assertEqual(report["selection"]["hidden_rows"], [6])
        self.assertIn("layout_policy_valid", report["summary"]["failed_checks"])

    def test_merged_cells_inside_range_are_rejected(self) -> None:
        def merge_data(workbook: Any) -> None:
            workbook["Заказы"].merge_cells("A5:A6")

        report = self.audit_modified(merge_data)
        self.assertFalse(report["summary"]["valid"])
        self.assertEqual(report["selection"]["merged_intersections"], ["A5:A6"])
        self.assertIn("layout_policy_valid", report["summary"]["failed_checks"])

    def test_wrong_amount_type_is_rejected(self) -> None:
        def break_amount(workbook: Any) -> None:
            workbook["Заказы"]["D5"] = "oops"

        report = self.audit_modified(break_amount)
        amount = report["schema"]["columns"]["amount"]
        self.assertEqual(amount["invalid_type_rows"], [5])
        self.assertFalse(report["summary"]["valid"])

    def test_duplicate_business_key_is_rejected(self) -> None:
        def duplicate_order(workbook: Any) -> None:
            workbook["Заказы"]["A6"] = workbook["Заказы"]["A5"].value

        report = self.audit_modified(duplicate_order)
        duplicates = report["schema"]["key"]["duplicate_rows"]
        self.assertEqual(duplicates[0]["row"], 6)
        self.assertEqual(duplicates[0]["duplicates_row"], 5)
        self.assertFalse(report["summary"]["valid"])

    def test_local_datetime_has_declared_source_timezone(self) -> None:
        ordered_at = self.report["schema"]["columns"]["ordered_at"]
        self.assertTrue(ordered_at["valid"])
        self.assertEqual(ordered_at["source_timezone"], "Europe/Moscow")
        self.assertEqual(
            ordered_at["localized_preview"][0],
            "2026-05-01T10:00:00+03:00",
        )

    def test_allowed_formula_without_cached_value_is_rejected(self) -> None:
        def keep_workbook_unchanged(workbook: Any) -> None:
            del workbook

        def include_formula_column(spec: dict[str, Any]) -> None:
            spec["boundary"]["range"] = "A4:G9"
            spec["columns"]["line_total"] = {"type": "number", "nullable": False}
            spec["formula_policy"] = "allow_with_cached_value"

        report = self.audit_modified(
            keep_workbook_unchanged,
            mutate_spec=include_formula_column,
        )
        self.assertFalse(report["formulas"]["valid"])
        self.assertEqual(
            report["formulas"]["missing_cached_values"],
            ["G5", "G6", "G7", "G8", "G9"],
        )
        self.assertIn("formula_policy_valid", report["summary"]["failed_checks"])

    def test_pandas_uses_the_same_explicit_selection(self) -> None:
        self.assertTrue(self.report["pandas"]["valid"])
        self.assertEqual(
            self.report["pandas"]["columns"],
            list(json.loads(SPEC_PATH.read_text())["columns"]),
        )

    def test_invalid_spec_is_rejected_before_workbook_audit(self) -> None:
        spec = json.loads(SPEC_PATH.read_text(encoding="utf-8"))
        spec["boundary"]["range"] = "A4:E9"
        with TemporaryDirectory() as directory:
            path = Path(directory) / "spec.json"
            path.write_text(json.dumps(spec), encoding="utf-8")
            with self.assertRaisesRegex(AUDITOR.ExcelAuditError, "range width"):
                AUDITOR.audit_workbook(VALID, path)

    def test_cli_returns_controlled_error_for_corrupt_workbook(self) -> None:
        with TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "broken.xlsx"
            workbook_path.write_bytes(b"not an xlsx file")
            result = subprocess.run(
                [
                    sys.executable,
                    ARTIFACT,
                    "--input",
                    workbook_path,
                    "--spec",
                    SPEC_PATH,
                ],
                check=False,
                capture_output=True,
                text=True,
            )
        self.assertEqual(result.returncode, 2)
        self.assertNotIn("Traceback", result.stderr)
        self.assertIn("cannot open workbook", json.loads(result.stdout)["error"])

    def test_cli_is_a_quality_gate(self) -> None:
        result = subprocess.run(
            [sys.executable, ARTIFACT, "--input", SHIFTED, "--spec", SPEC_PATH],
            check=False,
            capture_output=True,
            text=True,
        )
        self.assertEqual(result.returncode, 1, result.stderr)
        self.assertFalse(json.loads(result.stdout)["summary"]["valid"])


if __name__ == "__main__":
    unittest.main()
