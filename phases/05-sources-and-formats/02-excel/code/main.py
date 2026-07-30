import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT.parent / "data"
WORKBOOK = DATA / "tiny" / "orders_report.xlsx"
SPEC = json.loads((DATA / "excel_spec.json").read_text(encoding="utf-8"))

formula_book = load_workbook(WORKBOOK, data_only=False)
cached_book = load_workbook(WORKBOOK, data_only=True)
sheet = formula_book[SPEC["sheet"]]
cell_range = SPEC["boundary"]["range"]
min_col, min_row, max_col, max_row = range_boundaries(cell_range)

print("Активный лист:", formula_book.active.title)
print("Выбранный лист и диапазон:", SPEC["sheet"], cell_range)
print("Формула G5:", sheet["G5"].value)
print("Сохранённое значение G5:", cached_book[SPEC["sheet"]]["G5"].value)

hidden_in_range = [
    get_column_letter(column)
    for column in range(min_col, max_col + 1)
    if sheet.column_dimensions[get_column_letter(column)].hidden
]
spill_rows = [
    row
    for row in range(max_row + 1, sheet.max_row + 1)
    if any(
        sheet.cell(row=row, column=column).value is not None
        for column in range(min_col, max_col + 1)
    )
]
print("Скрытые столбцы внутри диапазона:", hidden_in_range)
print("Непустые строки ниже фиксированной границы:", spill_rows)

frame = pd.read_excel(
    WORKBOOK,
    sheet_name=SPEC["sheet"],
    header=min_row - 1,
    usecols=f"{get_column_letter(min_col)}:{get_column_letter(max_col)}",
    nrows=max_row - min_row,
    dtype=object,
    engine="openpyxl",
    keep_default_na=False,
)
print("Табличная схема:", frame.columns.tolist())
print("Строк:", len(frame), "Сумма amount:", frame["amount"].sum())
print(
    "Часовой пояс ordered_at:",
    SPEC["columns"]["ordered_at"]["source_timezone"],
)
