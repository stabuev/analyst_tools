from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import sqlite3
import zipfile
from datetime import UTC, datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any
from xml.etree import ElementTree

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ENCODING = "cp1251"
COLUMNS = ["order_id", "user_id", "ordered_at", "amount", "currency", "comment"]
VALID_ROWS = [
    ["O2001", "U001", "2026-05-01T10:00:00+03:00", "1 200,50", "RUB", "первый заказ"],
    ["O2002", "U002", "2026-05-02T08:30:00Z", "950,00", "RUB", "скидка; май"],
    ["O2003", "U003", "2026-05-03T11:15:00+01:00", "25,99", "EUR", "NULL"],
    ["O2004", "U004", "2026-05-04T17:45:00+05:00", "0,00", "RUB", ""],
    ["O2005", "U005", "2026-05-05T09:10:00Z", "1 050,10", "RUB", "повторный клиент"],
]
EXCEL_ROWS = [
    ["O2101", "U001", datetime(2026, 5, 1, 10, 0), 1200.5, "RUB", 1],
    ["O2102", "U002", datetime(2026, 5, 2, 8, 30), 475.0, "RUB", 2],
    ["O2103", "U003", datetime(2026, 5, 3, 11, 15), 25.99, "EUR", 1],
    ["O2104", "U004", datetime(2026, 5, 4, 17, 45), 0.0, "RUB", 1],
    ["O2105", "U005", datetime(2026, 5, 5, 9, 10), 350.1, "RUB", 3],
]
EVENTS = {
    "exported_at": "2026-05-06T00:00:00Z",
    "events": [
        {
            "event_id": "E5001",
            "user": {"id": "U001"},
            "occurred_at": "2026-05-01T10:00:00Z",
            "context": {"device": {"os": "ios"}, "screen": "checkout"},
            "items": [
                {"product_id": "P01", "quantity": 1, "price": 1000.0},
                {"product_id": "P02", "quantity": 1, "price": 200.5},
            ],
        },
        {
            "event_id": "E5002",
            "user": {"id": "U002"},
            "occurred_at": "2026-05-02T08:30:00Z",
            "context": {"device": {"os": "android"}, "screen": "catalog"},
            "items": [],
        },
        {
            "event_id": "E5003",
            "user": {"id": "U003"},
            "occurred_at": "2026-05-03T11:15:00Z",
            "context": {"device": {"os": None}, "screen": "checkout"},
            "items": [{"product_id": "P03", "quantity": 2, "price": 12.995}],
        },
    ],
}
API_PAGES = [
    {
        "page": 1,
        "items": [
            {
                "order_id": "O2301",
                "user_id": "U001",
                "ordered_at": "2026-05-01T10:00:00Z",
                "amount": 1200.5,
                "currency": "RUB",
                "comment": "first",
            },
            {
                "order_id": "O2302",
                "user_id": "U002",
                "ordered_at": "2026-05-02T08:30:00Z",
                "amount": 950.0,
                "currency": "RUB",
                "comment": None,
            },
        ],
        "next": "https://api.example.test/orders?page=2",
    },
    {
        "page": 2,
        "items": [
            {
                "order_id": "O2303",
                "user_id": "U003",
                "ordered_at": "2026-05-03T11:15:00Z",
                "amount": 25.99,
                "currency": "EUR",
                "comment": "promo",
            },
            {
                "order_id": "O2304",
                "user_id": "U004",
                "ordered_at": "2026-05-04T17:45:00Z",
                "amount": 0.0,
                "currency": "RUB",
                "comment": None,
            },
        ],
        "next": "https://api.example.test/orders?page=3",
    },
    {
        "page": 3,
        "items": [
            {
                "order_id": "O2305",
                "user_id": "U005",
                "ordered_at": "2026-05-05T09:10:00Z",
                "amount": 1050.1,
                "currency": "RUB",
                "comment": "repeat",
            }
        ],
        "next": None,
    },
]
TYPED_ORDERS = [
    ["order_id", "user_id", "ordered_at", "amount", "currency", "comment"],
    ["O2401", "U001", "2026-05-01T10:00:00Z", "1200.50", "RUB", "first"],
    ["O2402", "U002", "2026-05-02T08:30:00Z", "950.00", "RUB", ""],
    ["O2403", "U003", "2026-05-03T11:15:00Z", "25.99", "EUR", "promo"],
    ["O2404", "U004", "2026-05-04T17:45:00Z", "0.00", "RUB", ""],
    ["O2405", "U005", "2026-05-05T09:10:00Z", "1050.10", "RUB", "repeat"],
]


def render_csv(rows: list[list[str]]) -> str:
    output = io.StringIO(newline="")
    writer = csv.writer(
        output,
        delimiter=";",
        quotechar='"',
        doublequote=True,
        lineterminator="\n",
    )
    writer.writerows(rows)
    return output.getvalue()


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def write_bytes(path: Path, content: str) -> None:
    path.write_bytes(content.encode(ENCODING))


def write_json(path: Path, value: Any) -> None:
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def build_workbook(*, shifted: bool) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.created = datetime(2026, 5, 6, tzinfo=UTC)
    workbook.properties.modified = datetime(2026, 5, 6, tzinfo=UTC)

    instructions = workbook.create_sheet("Инструкция")
    instructions["A1"] = "Выгрузка заказов"
    instructions["A2"] = "Рабочие данные находятся на листе «Заказы»."

    sheet = workbook.create_sheet("Заказы")
    sheet.merge_cells("A1:G1")
    sheet["A1"] = "Отчет по заказам"
    sheet["A1"].font = Font(bold=True)
    sheet["A2"] = "Период"
    sheet["B2"] = "2026-05-01 — 2026-05-05"
    header_row = 5 if shifted else 4
    if shifted:
        sheet["A4"] = "Сформировано повторно: структура сдвинута"
    headers = [
        "order_id",
        "user_id",
        "ordered_at",
        "amount",
        "currency",
        "item_count",
        "line_total",
    ]
    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column, value=value)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for offset, row in enumerate(EXCEL_ROWS, start=1):
        row_number = header_row + offset
        for column, value in enumerate(row, start=1):
            sheet.cell(row=row_number, column=column, value=value)
        sheet.cell(row=row_number, column=7, value=f"=D{row_number}*F{row_number}")
        sheet.cell(row=row_number, column=3).number_format = "yyyy-mm-dd hh:mm"
        sheet.cell(row=row_number, column=4).number_format = "#,##0.00"
        sheet.cell(row=row_number, column=7).number_format = "#,##0.00"
    sheet.column_dimensions["H"].hidden = True
    sheet["H1"] = "internal_note"

    dictionary = workbook.create_sheet("Справочник")
    dictionary.append(["currency", "description"])
    dictionary.append(["RUB", "Российский рубль"])
    dictionary.append(["EUR", "Евро"])
    return workbook


def save_workbook_deterministic(workbook: Workbook, path: Path) -> None:
    with TemporaryDirectory() as directory:
        raw_path = Path(directory) / "raw.xlsx"
        workbook.save(raw_path)
        with (
            zipfile.ZipFile(raw_path, "r") as source,
            zipfile.ZipFile(
                path,
                "w",
                compression=zipfile.ZIP_DEFLATED,
                compresslevel=9,
            ) as target,
        ):
            for name in sorted(source.namelist()):
                info = zipfile.ZipInfo(name, date_time=(2026, 5, 6, 0, 0, 0))
                info.compress_type = zipfile.ZIP_DEFLATED
                info.external_attr = 0o600 << 16
                content = source.read(name)
                if name == "docProps/core.xml":
                    namespaces = {
                        "cp": "http://schemas.openxmlformats.org/package/2006/metadata/core-properties",
                        "dc": "http://purl.org/dc/elements/1.1/",
                        "dcterms": "http://purl.org/dc/terms/",
                        "xsi": "http://www.w3.org/2001/XMLSchema-instance",
                    }
                    for prefix, uri in namespaces.items():
                        ElementTree.register_namespace(prefix, uri)
                    root = ElementTree.fromstring(content)
                    for field in ("created", "modified"):
                        node = root.find(f"{{{namespaces['dcterms']}}}{field}")
                        if node is not None:
                            node.text = "2026-05-06T00:00:00Z"
                    content = ElementTree.tostring(root, encoding="utf-8")
                target.writestr(info, content)


def write_sqlite(path: Path) -> None:
    path.unlink(missing_ok=True)
    connection = sqlite3.connect(path)
    try:
        connection.executescript(
            """
            PRAGMA page_size = 4096;
            CREATE TABLE users (
                user_id TEXT PRIMARY KEY,
                segment TEXT NOT NULL
            );
            CREATE TABLE orders (
                order_id TEXT PRIMARY KEY,
                user_id TEXT NOT NULL,
                ordered_at TEXT NOT NULL,
                amount REAL NOT NULL,
                status TEXT NOT NULL
            );
            """
        )
        connection.executemany(
            "INSERT INTO users VALUES (?, ?)",
            [
                ("U001", "new"),
                ("U002", "new"),
                ("U003", "returning"),
                ("U004", "returning"),
                ("U005", "vip"),
            ],
        )
        connection.executemany(
            "INSERT INTO orders VALUES (?, ?, ?, ?, ?)",
            [
                ("O2501", "U001", "2026-05-01T10:00:00Z", 1200.5, "paid"),
                ("O2502", "U002", "2026-05-02T08:30:00Z", 950.0, "paid"),
                ("O2503", "U003", "2026-05-03T11:15:00Z", 25.99, "cancelled"),
                ("O2504", "U004", "2026-05-04T17:45:00Z", 0.0, "refunded"),
                ("O2505", "U005", "2026-05-05T09:10:00Z", 1050.1, "paid"),
            ],
        )
        connection.commit()
        connection.execute("VACUUM")
    finally:
        connection.close()


def generate(output_dir: Path) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    valid_path = output_dir / "orders_semicolon_cp1251.csv"
    broken_path = output_dir / "orders_broken_cp1251.csv"

    write_bytes(valid_path, render_csv([COLUMNS, *VALID_ROWS]))

    broken = (
        render_csv([COLUMNS, VALID_ROWS[0]])
        + "O2002;U002;2026-05-02T08:30:00Z;950,00;RUB;скидка; май\n"
        + render_csv([VALID_ROWS[2], VALID_ROWS[3]])
    )
    write_bytes(broken_path, broken)

    workbook_path = output_dir / "orders_report.xlsx"
    shifted_workbook_path = output_dir / "orders_report_shifted.xlsx"
    save_workbook_deterministic(build_workbook(shifted=False), workbook_path)
    save_workbook_deterministic(build_workbook(shifted=True), shifted_workbook_path)

    events_path = output_dir / "events_nested.json"
    drift_path = output_dir / "events_schema_drift.json"
    write_json(events_path, EVENTS)
    drift = json.loads(json.dumps(EVENTS))
    drift["events"][0]["context"]["app_version"] = "6.1.0"
    drift["events"][2]["items"][0]["price"] = "12.995"
    write_json(drift_path, drift)

    http_body_path = output_dir / "http_orders.json"
    write_json(
        http_body_path,
        {
            "next": None,
            "orders": [
                {"order_id": "O2201", "amount": 1200.5},
                {"order_id": "O2202", "amount": 950.0},
            ],
        },
    )

    api_paths = []
    for page in API_PAGES:
        page_path = output_dir / f"api_page_{page['page']}.json"
        write_json(page_path, page)
        api_paths.append(page_path)

    html_path = output_dir / "orders.html"
    changed_html_path = output_dir / "orders_changed.html"
    html_path.write_text(
        """<!doctype html>
<html lang="ru"><body><section data-orders>
  <article data-order-card data-order-id="O2601">
    <span data-field="user">U001</span><span data-field="amount">1200.50</span>
  </article>
  <article data-order-card data-order-id="O2602">
    <span data-field="user">U002</span><span data-field="amount">950.00</span>
  </article>
</section></body></html>
""",
        encoding="utf-8",
    )
    changed_html_path.write_text(
        """<!doctype html>
<html lang="ru"><body><section data-orders>
  <article data-order-card data-order-id="O2601">
    <span data-field="user">U001</span><span data-field="amount">1200.50</span>
  </article>
  <article data-order-card data-order-id="O2602">
    <span data-field="user">U002</span><strong data-field="total">950.00</strong>
  </article>
</section></body></html>
""",
        encoding="utf-8",
    )

    database_path = output_dir / "analytics.sqlite"
    write_sqlite(database_path)

    typed_csv_path = output_dir / "orders_typed.csv"
    typed_csv_path.write_text(render_csv(TYPED_ORDERS).replace(";", ","), encoding="utf-8")

    generated = {
        valid_path.name: {"kind": "csv", "valid_by_contract": True},
        broken_path.name: {"kind": "csv", "valid_by_contract": False},
        workbook_path.name: {"kind": "xlsx", "valid_by_contract": True},
        shifted_workbook_path.name: {"kind": "xlsx", "valid_by_contract": False},
        events_path.name: {"kind": "json", "valid_by_contract": True},
        drift_path.name: {"kind": "json", "valid_by_contract": False},
        http_body_path.name: {"kind": "http-body", "valid_by_contract": True},
        **{path.name: {"kind": "api-page", "valid_by_contract": True} for path in api_paths},
        html_path.name: {"kind": "html", "valid_by_contract": True},
        changed_html_path.name: {"kind": "html", "valid_by_contract": False},
        database_path.name: {"kind": "sqlite", "valid_by_contract": True},
        typed_csv_path.name: {"kind": "csv", "valid_by_contract": True},
    }
    manifest = {
        "version": "1.0.0",
        "generated_by": "generate_data.py",
        "files": {},
    }
    for name, metadata in generated.items():
        path = output_dir / name
        manifest["files"][name] = {
            **metadata,
            "bytes": path.stat().st_size,
            "sha256": sha256(path),
        }
    manifest["files"][valid_path.name].update({"encoding": ENCODING, "data_rows": len(VALID_ROWS)})
    manifest["files"][broken_path.name].update({"encoding": ENCODING, "data_rows": 4})
    (output_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    return manifest


def check_committed(output_dir: Path) -> None:
    with TemporaryDirectory() as directory:
        generated_dir = Path(directory)
        generate(generated_dir)
        expected_names = {path.name for path in generated_dir.iterdir() if path.is_file()}
        actual_names = {path.name for path in output_dir.iterdir() if path.is_file()}
        if actual_names != expected_names:
            raise ValueError(
                f"tracked fixture names differ: expected {sorted(expected_names)}, "
                f"got {sorted(actual_names)}"
            )
        for name in sorted(expected_names):
            if (output_dir / name).read_bytes() != (generated_dir / name).read_bytes():
                raise ValueError(f"tracked fixture is stale: {name}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate deterministic phase 05 fixtures")
    parser.add_argument("--output", type=Path)
    parser.add_argument("--check", action="store_true")
    args = parser.parse_args()

    output = args.output or Path(__file__).parent / "tiny"
    if args.check:
        check_committed(output)
        print(f"Phase 05 fixtures are up to date in {output}")
        return

    manifest = generate(output)
    print(f"Generated {len(manifest['files'])} phase fixtures in {output}")


if __name__ == "__main__":
    main()
