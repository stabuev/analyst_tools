from __future__ import annotations

import csv
import importlib.util
import json
import subprocess
import sys
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
ARTIFACT = ROOT / "outputs" / "stakeholder_workbook_builder.py"
CODE = ROOT / "code" / "main.py"
SPEC = importlib.util.spec_from_file_location("stakeholder_workbook_builder", ARTIFACT)
if SPEC is None or SPEC.loader is None:
    raise RuntimeError(f"cannot load {ARTIFACT}")
BUILDER = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = BUILDER
SPEC.loader.exec_module(BUILDER)


def read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: dict) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8", newline="") as source:
        return list(csv.DictReader(source))


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as target:
        writer = csv.DictWriter(target, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def build_sample(root: Path):
    paths = BUILDER.write_sample_inputs(root / "inputs")
    result = BUILDER.build_stakeholder_workbook(
        spec_path=paths["spec_path"],
        metrics_path=paths["metrics_path"],
        evidence_path=paths["evidence_path"],
        memo_audit_path=paths["memo_audit_path"],
        output_dir=root / "workbook",
    )
    return paths, result


def check_by_id(audit: dict, check_id: str) -> dict:
    return next(check for check in audit["checks"] if check["id"] == check_id)


class StakeholderWorkbookBuilderTest(unittest.TestCase):
    def test_sample_workbook_is_valid_and_writes_package_files(self) -> None:
        with TemporaryDirectory() as directory:
            _paths, result = build_sample(Path(directory))

            self.assertTrue(result.audit["valid"])
            self.assertEqual(result.audit["readiness_status"], "ready")
            self.assertEqual(result.audit["summary"]["blocking_errors"], [])
            for path in [
                result.workbook_path,
                result.audit_path,
                result.dictionary_path,
                result.manifest_path,
            ]:
                self.assertTrue(path.is_file(), path)

    def test_workbook_contains_required_sheets_tables_and_freeze_panes(self) -> None:
        with TemporaryDirectory() as directory:
            _paths, result = build_sample(Path(directory))
            workbook = load_workbook(result.workbook_path, data_only=False)

            self.assertEqual(workbook.sheetnames, BUILDER.REQUIRED_SHEETS)
            self.assertEqual(set(BUILDER.workbook_tables(result.workbook_path)), {
                "MetricsTable",
                "EvidenceTable",
                "DictionaryTable",
                "ChecksTable",
            })
            for sheet in ["Metrics", "Evidence", "Data Dictionary", "Checks"]:
                self.assertEqual(workbook[sheet].freeze_panes, "A4")

    def test_summary_formulas_have_cached_values_matching_metric_source(self) -> None:
        with TemporaryDirectory() as directory:
            paths, result = build_sample(Path(directory))
            formulas = load_workbook(result.workbook_path, data_only=False)["Summary"]
            values = load_workbook(result.workbook_path, data_only=True)["Summary"]
            metrics = read_csv(paths["metrics_path"])
            totals = BUILDER.metric_totals(metrics)

            self.assertEqual(formulas["B10"].value, "=SUM(Metrics!C4:C6)")
            self.assertEqual(formulas["B11"].value, "=SUM(Metrics!D4:D6)")
            self.assertEqual(formulas["B12"].value, '=COUNTIF(Metrics!F4:F6,"breached")')
            self.assertAlmostEqual(values["B10"].value, totals["current_total"])
            self.assertAlmostEqual(values["B11"].value, totals["baseline_total"])
            self.assertEqual(values["B12"].value, totals["breached_count"])

    def test_metrics_and_evidence_rows_preserve_source_order(self) -> None:
        with TemporaryDirectory() as directory:
            paths, result = build_sample(Path(directory))
            workbook = load_workbook(result.workbook_path, data_only=True)

            metric_ids = [
                workbook["Metrics"].cell(row=row_number, column=1).value
                for row_number in range(4, 7)
            ]
            evidence_ids = [
                workbook["Evidence"].cell(row=row_number, column=2).value
                for row_number in range(4, 8)
            ]
            self.assertEqual(metric_ids, [row["metric_id"] for row in read_csv(paths["metrics_path"])])
            self.assertEqual(evidence_ids, [row["evidence_id"] for row in read_csv(paths["evidence_path"])])

    def test_data_dictionary_covers_visible_metric_and_evidence_columns(self) -> None:
        with TemporaryDirectory() as directory:
            _paths, result = build_sample(Path(directory))
            workbook = load_workbook(result.workbook_path, data_only=True)
            rows = {
                (
                    workbook["Data Dictionary"].cell(row=row_number, column=1).value,
                    workbook["Data Dictionary"].cell(row=row_number, column=2).value,
                )
                for row_number in range(4, workbook["Data Dictionary"].max_row + 1)
            }

            for column in BUILDER.REQUIRED_METRIC_COLUMNS:
                self.assertIn(("Metrics", column), rows)
            for column in BUILDER.REQUIRED_EVIDENCE_COLUMNS:
                self.assertIn(("Evidence", column), rows)

    def test_blocked_upstream_memo_audit_blocks_workbook(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            paths = BUILDER.write_sample_inputs(root / "inputs")
            memo_audit = read_json(paths["memo_audit_path"])
            memo_audit["valid"] = False
            memo_audit["summary"]["blocking_errors"] = ["no_unsupported_overclaim_wording"]
            write_json(paths["memo_audit_path"], memo_audit)

            result = BUILDER.build_stakeholder_workbook(**paths, output_dir=root / "workbook")

            self.assertFalse(result.audit["valid"])
            self.assertFalse(check_by_id(result.audit, "upstream_memo_audit_is_valid")["valid"])

    def test_missing_metric_column_blocks_input_contract(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            paths = BUILDER.write_sample_inputs(root / "inputs")
            rows = read_csv(paths["metrics_path"])
            for row in rows:
                row.pop("threshold")
            write_csv(paths["metrics_path"], rows, [column for column in BUILDER.REQUIRED_METRIC_COLUMNS if column != "threshold"])

            result = BUILDER.build_stakeholder_workbook(**paths, output_dir=root / "workbook")

            self.assertFalse(result.audit["valid"])
            metric_check = check_by_id(result.audit, "metric_summary_has_required_columns")
            self.assertFalse(metric_check["valid"])
            self.assertEqual(metric_check["observed"], ["threshold"])

    def test_unknown_metric_status_is_rejected(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            paths = BUILDER.write_sample_inputs(root / "inputs")
            rows = read_csv(paths["metrics_path"])
            rows[0]["status"] = "ship_anyway"
            write_csv(paths["metrics_path"], rows, BUILDER.REQUIRED_METRIC_COLUMNS)

            result = BUILDER.build_stakeholder_workbook(**paths, output_dir=root / "workbook")

            self.assertFalse(result.audit["valid"])
            status_check = check_by_id(result.audit, "metric_statuses_are_known")
            self.assertFalse(status_check["valid"])
            self.assertEqual(status_check["observed"], ["support_ticket_rate_7d"])

    def test_sensitive_dictionary_column_blocks_workbook_publication(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            paths = BUILDER.write_sample_inputs(root / "inputs")
            spec = read_json(paths["spec_path"])
            spec["data_dictionary"][0]["sensitive"] = True
            write_json(paths["spec_path"], spec)

            result = BUILDER.build_stakeholder_workbook(**paths, output_dir=root / "workbook")

            self.assertFalse(result.audit["valid"])
            sensitive_check = check_by_id(result.audit, "no_sensitive_columns_in_workbook")
            self.assertFalse(sensitive_check["valid"])
            self.assertEqual(sensitive_check["observed"], ["Metrics.metric_id"])

    def test_workbook_audit_detects_tampered_summary_formula(self) -> None:
        with TemporaryDirectory() as directory:
            paths, result = build_sample(Path(directory))
            workbook = load_workbook(result.workbook_path)
            workbook["Summary"]["B10"] = 0
            workbook.save(result.workbook_path)

            audit = BUILDER.audit_workbook(
                workbook_path=result.workbook_path,
                spec=read_json(paths["spec_path"]),
                metrics=read_csv(paths["metrics_path"]),
                evidence=read_csv(paths["evidence_path"]),
                initial_checks=[],
            )

            self.assertFalse(audit["valid"])
            self.assertFalse(check_by_id(audit, "summary_formulas_are_present")["valid"])

    def test_manifest_hashes_inputs_and_outputs(self) -> None:
        with TemporaryDirectory() as directory:
            _paths, result = build_sample(Path(directory))
            manifest = read_json(result.manifest_path)

            self.assertEqual(manifest["hash_algorithm"], "sha256")
            self.assertEqual(
                set(manifest["inputs"]),
                {"workbook_spec", "metric_summary", "claim_evidence_matrix", "memo_audit"},
            )
            self.assertEqual(
                set(manifest["outputs"]),
                {"stakeholder_workbook", "workbook_audit", "data_dictionary"},
            )
            all_hashes = [
                item["sha256"]
                for section in ("inputs", "outputs")
                for item in manifest[section].values()
            ]
            self.assertTrue(all(len(value) == 64 for value in all_hashes))

    def test_cli_write_example_builds_workbook_package(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            proc = subprocess.run(
                [
                    sys.executable,
                    str(ARTIFACT),
                    "--write-example",
                    str(root / "inputs"),
                    "--output-dir",
                    str(root / "workbook"),
                ],
                check=False,
                capture_output=True,
                text=True,
            )

            self.assertEqual(proc.returncode, 0, proc.stderr)
            report = json.loads(proc.stdout)
            self.assertTrue(report["valid"])
            self.assertTrue((root / "workbook" / "stakeholder_workbook.xlsx").is_file())

    def test_cli_returns_nonzero_for_invalid_workbook_when_requested(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            paths = BUILDER.write_sample_inputs(root / "inputs")
            spec = read_json(paths["spec_path"])
            spec["data_dictionary"][0]["sensitive"] = True
            write_json(paths["spec_path"], spec)
            proc = subprocess.run(
                [
                    sys.executable,
                    str(ARTIFACT),
                    "--spec",
                    str(paths["spec_path"]),
                    "--metrics",
                    str(paths["metrics_path"]),
                    "--evidence",
                    str(paths["evidence_path"]),
                    "--memo-audit",
                    str(paths["memo_audit_path"]),
                    "--output-dir",
                    str(root / "workbook"),
                    "--fail-on-invalid",
                ],
                check=False,
                capture_output=True,
                text=True,
            )

            self.assertEqual(proc.returncode, 2, proc.stderr)
            self.assertFalse(json.loads(proc.stdout)["valid"])

    def test_code_example_runs_without_external_files(self) -> None:
        proc = subprocess.run(
            [sys.executable, str(CODE)],
            check=False,
            capture_output=True,
            text=True,
        )

        self.assertEqual(proc.returncode, 0, proc.stderr)
        payload = json.loads(proc.stdout)
        self.assertTrue(payload["valid"])
        self.assertEqual(payload["files"], [
            "stakeholder_workbook.xlsx",
            "workbook_audit.json",
            "data_dictionary.csv",
            "manifest.json",
        ])


if __name__ == "__main__":
    unittest.main()
