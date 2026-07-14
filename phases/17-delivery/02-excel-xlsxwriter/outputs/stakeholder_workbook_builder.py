from __future__ import annotations

import argparse
import csv
import hashlib
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import xlsxwriter
from openpyxl import load_workbook


BUILDER_VERSION = "1.0.0"
REQUIRED_SHEETS = ["Summary", "Metrics", "Evidence", "Data Dictionary", "Checks"]
REQUIRED_TABLES = {
    "Metrics": "MetricsTable",
    "Evidence": "EvidenceTable",
    "Data Dictionary": "DictionaryTable",
    "Checks": "ChecksTable",
}
REQUIRED_METRIC_COLUMNS = [
    "metric_id",
    "label",
    "current",
    "baseline",
    "threshold",
    "status",
    "owner",
]
REQUIRED_EVIDENCE_COLUMNS = [
    "claim_id",
    "evidence_id",
    "metric_id",
    "quality_status",
    "decision_impact",
]
REQUIRED_SPEC_FIELDS = {
    "workbook_id",
    "title",
    "audience",
    "decision_owner",
    "source_memo_id",
    "decision_status",
    "readiness_status",
    "freshness",
    "data_dictionary",
}
ALLOWED_METRIC_STATUSES = {"ok", "watch", "breached"}
ALLOWED_QUALITY_STATUSES = {"pass", "warn", "block", "missing"}


@dataclass(frozen=True)
class WorkbookBuildResult:
    output_dir: Path
    workbook_path: Path
    audit_path: Path
    dictionary_path: Path
    manifest_path: Path
    audit: dict[str, Any]


def read_json(path: str | Path) -> dict[str, Any]:
    return json.loads(Path(path).read_text(encoding="utf-8"))


def write_json(path: str | Path, payload: Any) -> None:
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def read_csv(path: str | Path) -> list[dict[str, str]]:
    with Path(path).open(encoding="utf-8", newline="") as source:
        return list(csv.DictReader(source))


def write_csv(path: str | Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    with destination.open("w", encoding="utf-8", newline="") as target:
        writer = csv.DictWriter(target, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def sha256_file(path: str | Path) -> str:
    return hashlib.sha256(Path(path).read_bytes()).hexdigest()


def as_number(value: str, *, field: str) -> float:
    try:
        return float(value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"{field} must be numeric: {value!r}") from error


def sample_workbook_spec() -> dict[str, Any]:
    return {
        "workbook_id": "trial-onboarding-stakeholder-workbook",
        "title": "Stakeholder workbook: onboarding rollout risk",
        "audience": "Growth weekly decision review",
        "decision_owner": "Head of Growth",
        "source_memo_id": "trial-onboarding-risk-memo",
        "decision_status": "pause_rollout",
        "readiness_status": "ready_with_warnings",
        "freshness": "2026-05-21",
        "source_paths": {
            "memo": "executive_memo.md",
            "audit": "memo_audit.json",
            "evidence": "claim_evidence_matrix.csv",
        },
        "data_dictionary": [
            {
                "sheet": "Metrics",
                "column": "metric_id",
                "description": "Stable metric identifier used in memo evidence.",
                "source": "metric_summary.csv",
                "expected_type": "string",
                "sensitive": False,
            },
            {
                "sheet": "Metrics",
                "column": "label",
                "description": "Readable metric label for stakeholder review.",
                "source": "metric_summary.csv",
                "expected_type": "string",
                "sensitive": False,
            },
            {
                "sheet": "Metrics",
                "column": "current",
                "description": "Current complete-window metric value.",
                "source": "metric_summary.csv",
                "expected_type": "number",
                "sensitive": False,
            },
            {
                "sheet": "Metrics",
                "column": "baseline",
                "description": "Reference metric value used by the memo.",
                "source": "metric_summary.csv",
                "expected_type": "number",
                "sensitive": False,
            },
            {
                "sheet": "Metrics",
                "column": "threshold",
                "description": "Decision threshold or guardrail boundary.",
                "source": "metric_summary.csv",
                "expected_type": "number",
                "sensitive": False,
            },
            {
                "sheet": "Metrics",
                "column": "status",
                "description": "Stakeholder status: ok, watch or breached.",
                "source": "metric_summary.csv",
                "expected_type": "category",
                "sensitive": False,
            },
            {
                "sheet": "Metrics",
                "column": "owner",
                "description": "Owner responsible for follow-up on the metric.",
                "source": "metric_summary.csv",
                "expected_type": "string",
                "sensitive": False,
            },
            {
                "sheet": "Evidence",
                "column": "claim_id",
                "description": "Claim identifier inherited from the decision memo.",
                "source": "claim_evidence_matrix.csv",
                "expected_type": "string",
                "sensitive": False,
            },
            {
                "sheet": "Evidence",
                "column": "evidence_id",
                "description": "Evidence identifier cited by the memo.",
                "source": "claim_evidence_matrix.csv",
                "expected_type": "string",
                "sensitive": False,
            },
            {
                "sheet": "Evidence",
                "column": "metric_id",
                "description": "Metric or context identifier for the evidence row.",
                "source": "claim_evidence_matrix.csv",
                "expected_type": "string",
                "sensitive": False,
            },
            {
                "sheet": "Evidence",
                "column": "quality_status",
                "description": "Evidence quality inherited from the memo audit.",
                "source": "claim_evidence_matrix.csv",
                "expected_type": "category",
                "sensitive": False,
            },
            {
                "sheet": "Evidence",
                "column": "decision_impact",
                "description": "How the evidence affects the workbook recommendation.",
                "source": "claim_evidence_matrix.csv",
                "expected_type": "category",
                "sensitive": False,
            },
        ],
    }


def sample_metric_rows() -> list[dict[str, str]]:
    return [
        {
            "metric_id": "support_ticket_rate_7d",
            "label": "Support ticket rate, 7d",
            "current": "0.018",
            "baseline": "0.011",
            "threshold": "0.010",
            "status": "breached",
            "owner": "Support analytics",
        },
        {
            "metric_id": "subscription_cancel_rate_14d",
            "label": "Subscription cancellation rate, 14d",
            "current": "0.031",
            "baseline": "0.022",
            "threshold": "0.024",
            "status": "breached",
            "owner": "Growth analytics",
        },
        {
            "metric_id": "support_reason_coverage",
            "label": "Support reason coverage",
            "current": "0.740",
            "baseline": "0.910",
            "threshold": "0.900",
            "status": "watch",
            "owner": "Support analytics",
        },
    ]


def sample_evidence_rows() -> list[dict[str, str]]:
    return [
        {
            "claim_id": "guardrails-above-threshold",
            "evidence_id": "support-ticket-rate",
            "metric_id": "support_ticket_rate_7d",
            "quality_status": "pass",
            "decision_impact": "usable",
        },
        {
            "claim_id": "guardrails-above-threshold",
            "evidence_id": "cancel-rate",
            "metric_id": "subscription_cancel_rate_14d",
            "quality_status": "pass",
            "decision_impact": "usable",
        },
        {
            "claim_id": "quality-gates-usable",
            "evidence_id": "support-reason-coverage",
            "metric_id": "support_reason_coverage",
            "quality_status": "warn",
            "decision_impact": "usable_with_disclosure",
        },
        {
            "claim_id": "calendar-context-only",
            "evidence_id": "release-calendar",
            "metric_id": "__context__",
            "quality_status": "pass",
            "decision_impact": "context_only",
        },
    ]


def sample_memo_audit() -> dict[str, Any]:
    return {
        "valid": True,
        "readiness_status": "ready_with_warnings",
        "memo_id": "trial-onboarding-risk-memo",
        "recommended_decision": "pause_rollout",
        "summary": {
            "blocking_errors": [],
            "warnings": [
                "quality_gate_warnings_are_visible",
                "evidence_quality_warnings_are_disclosed",
            ],
        },
    }


def write_sample_inputs(root: str | Path) -> dict[str, Path]:
    root_path = Path(root)
    root_path.mkdir(parents=True, exist_ok=True)
    spec_path = root_path / "workbook_spec.json"
    metrics_path = root_path / "metric_summary.csv"
    evidence_path = root_path / "claim_evidence_matrix.csv"
    memo_audit_path = root_path / "memo_audit.json"
    write_json(spec_path, sample_workbook_spec())
    write_csv(metrics_path, sample_metric_rows(), REQUIRED_METRIC_COLUMNS)
    write_csv(evidence_path, sample_evidence_rows(), REQUIRED_EVIDENCE_COLUMNS)
    write_json(memo_audit_path, sample_memo_audit())
    return {
        "spec_path": spec_path,
        "metrics_path": metrics_path,
        "evidence_path": evidence_path,
        "memo_audit_path": memo_audit_path,
    }


def check(
    check_id: str,
    valid: bool,
    *,
    severity: str = "block",
    observed: Any = None,
    expected: Any = None,
    message: str = "",
) -> dict[str, Any]:
    return {
        "id": check_id,
        "valid": bool(valid),
        "severity": severity,
        "observed": observed,
        "expected": expected,
        "message": message,
    }


def validate_inputs(
    spec: dict[str, Any],
    metrics: list[dict[str, str]],
    evidence: list[dict[str, str]],
    memo_audit: dict[str, Any],
) -> list[dict[str, Any]]:
    checks: list[dict[str, Any]] = []
    missing_spec_fields = sorted(REQUIRED_SPEC_FIELDS - set(spec))
    checks.append(
        check(
            "spec_has_required_fields",
            not missing_spec_fields,
            observed=missing_spec_fields,
            expected=[],
            message="Workbook spec must name audience, owner, source memo and dictionary.",
        )
    )

    missing_metric_columns = sorted(
        {
            column
            for row in metrics
            for column in REQUIRED_METRIC_COLUMNS
            if column not in row
        }
    )
    checks.append(
        check(
            "metric_summary_has_required_columns",
            bool(metrics) and not missing_metric_columns,
            observed=missing_metric_columns,
            expected=[],
            message="Metric summary drives the stakeholder-facing Metrics sheet.",
        )
    )

    missing_evidence_columns = sorted(
        {
            column
            for row in evidence
            for column in REQUIRED_EVIDENCE_COLUMNS
            if column not in row
        }
    )
    checks.append(
        check(
            "evidence_matrix_has_required_columns",
            bool(evidence) and not missing_evidence_columns,
            observed=missing_evidence_columns,
            expected=[],
            message="Workbook evidence must preserve memo claim/evidence lineage.",
        )
    )

    invalid_metric_statuses = sorted(
        {
            row.get("metric_id", "<missing-id>")
            for row in metrics
            if row.get("status") not in ALLOWED_METRIC_STATUSES
        }
    )
    checks.append(
        check(
            "metric_statuses_are_known",
            not invalid_metric_statuses,
            observed=invalid_metric_statuses,
            expected=sorted(ALLOWED_METRIC_STATUSES),
            message="Unknown metric statuses make conditional formatting ambiguous.",
        )
    )

    invalid_quality_statuses = sorted(
        {
            row.get("evidence_id", "<missing-id>")
            for row in evidence
            if row.get("quality_status") not in ALLOWED_QUALITY_STATUSES
        }
    )
    checks.append(
        check(
            "evidence_quality_statuses_are_known",
            not invalid_quality_statuses,
            observed=invalid_quality_statuses,
            expected=sorted(ALLOWED_QUALITY_STATUSES),
            message="Unknown evidence quality cannot be explained on the Checks sheet.",
        )
    )

    numeric_errors: list[str] = []
    for row in metrics:
        for field in ("current", "baseline", "threshold"):
            try:
                as_number(row.get(field, ""), field=field)
            except ValueError:
                numeric_errors.append(f"{row.get('metric_id', '<missing-id>')}:{field}")
    checks.append(
        check(
            "metric_numeric_fields_are_numbers",
            not numeric_errors,
            observed=numeric_errors,
            expected=[],
            message="Workbook totals and formulas require numeric metric fields.",
        )
    )

    checks.append(
        check(
            "upstream_memo_audit_is_valid",
            bool(memo_audit.get("valid")),
            observed=memo_audit.get("summary", {}).get("blocking_errors", []),
            expected=[],
            message="Workbook cannot be published from a blocked decision memo.",
        )
    )

    dictionary_entries = spec.get("data_dictionary", [])
    dictionary_pairs = {
        (entry.get("sheet"), entry.get("column"))
        for entry in dictionary_entries
        if isinstance(entry, dict)
    }
    required_pairs = {
        ("Metrics", column) for column in REQUIRED_METRIC_COLUMNS
    } | {("Evidence", column) for column in REQUIRED_EVIDENCE_COLUMNS}
    missing_pairs = sorted(f"{sheet}.{column}" for sheet, column in required_pairs - dictionary_pairs)
    checks.append(
        check(
            "data_dictionary_covers_exported_columns",
            not missing_pairs,
            observed=missing_pairs,
            expected=[],
            message="Every stakeholder-visible data column needs a dictionary row.",
        )
    )

    sensitive_pairs = sorted(
        f"{entry.get('sheet')}.{entry.get('column')}"
        for entry in dictionary_entries
        if isinstance(entry, dict) and entry.get("sensitive") is True
    )
    checks.append(
        check(
            "no_sensitive_columns_in_workbook",
            not sensitive_pairs,
            observed=sensitive_pairs,
            expected=[],
            message="Stakeholder workbook must not expose columns marked sensitive.",
        )
    )
    return checks


def build_audit(checks: list[dict[str, Any]], *, workbook_id: str) -> dict[str, Any]:
    blockers = [item["id"] for item in checks if not item["valid"] and item["severity"] == "block"]
    warnings = [item["id"] for item in checks if not item["valid"] and item["severity"] == "warn"]
    return {
        "version": BUILDER_VERSION,
        "valid": not blockers,
        "workbook_id": workbook_id,
        "readiness_status": "blocked" if blockers else "ready",
        "summary": {
            "blocking_errors": blockers,
            "warnings": warnings,
            "check_count": len(checks),
        },
        "checks": checks,
    }


def metric_totals(metrics: list[dict[str, str]]) -> dict[str, float | int]:
    def numeric_or_zero(row: dict[str, str], field: str) -> float:
        try:
            return as_number(row.get(field, ""), field=field)
        except ValueError:
            return 0.0

    return {
        "current_total": round(sum(numeric_or_zero(row, "current") for row in metrics), 6),
        "baseline_total": round(
            sum(numeric_or_zero(row, "baseline") for row in metrics), 6
        ),
        "breached_count": sum(1 for row in metrics if row["status"] == "breached"),
    }


def write_table(
    worksheet: Any,
    *,
    start_row: int,
    start_col: int,
    table_name: str,
    columns: list[str],
    rows: list[dict[str, Any]],
    header_format: Any,
) -> tuple[int, int, int, int]:
    last_row = start_row + len(rows)
    last_col = start_col + len(columns) - 1
    for column_index, column in enumerate(columns):
        worksheet.write(start_row, start_col + column_index, column, header_format)
    for row_index, row in enumerate(rows, start=start_row + 1):
        for column_index, column in enumerate(columns):
            value = row.get(column, "")
            worksheet.write(row_index, start_col + column_index, value)
    worksheet.add_table(
        start_row,
        start_col,
        last_row,
        last_col,
        {
            "name": table_name,
            "columns": [{"header": column} for column in columns],
            "style": "Table Style Light 9",
        },
    )
    worksheet.freeze_panes(start_row + 1, 0)
    return start_row, start_col, last_row, last_col


def render_workbook(
    workbook_path: Path,
    spec: dict[str, Any],
    metrics: list[dict[str, str]],
    evidence: list[dict[str, str]],
    dictionary_rows: list[dict[str, Any]],
    initial_checks: list[dict[str, Any]],
) -> None:
    totals = metric_totals(metrics)
    workbook = xlsxwriter.Workbook(workbook_path)
    title_format = workbook.add_format({"bold": True, "font_size": 16, "font_color": "#1F2937"})
    label_format = workbook.add_format({"bold": True, "font_color": "#374151"})
    header_format = workbook.add_format(
        {
            "bold": True,
            "font_color": "#FFFFFF",
            "bg_color": "#1F4E79",
            "border": 1,
        }
    )
    number_format = workbook.add_format({"num_format": "0.0%"})
    integer_format = workbook.add_format({"num_format": "0"})
    warning_format = workbook.add_format({"bg_color": "#FFF2CC"})
    breach_format = workbook.add_format({"bg_color": "#F4CCCC"})

    summary = workbook.add_worksheet("Summary")
    summary.write("A1", spec["title"], title_format)
    summary.write("A3", "Audience", label_format)
    summary.write("B3", spec["audience"])
    summary.write("A4", "Decision owner", label_format)
    summary.write("B4", spec["decision_owner"])
    summary.write("A5", "Source memo", label_format)
    summary.write("B5", spec["source_memo_id"])
    summary.write("A6", "Decision status", label_format)
    summary.write("B6", spec["decision_status"])
    summary.write("A7", "Readiness", label_format)
    summary.write("B7", spec["readiness_status"])
    summary.write("A8", "Freshness", label_format)
    summary.write("B8", spec["freshness"])
    summary.write("A10", "Metric current total", label_format)
    summary.write_formula("B10", "=SUM(Metrics!C4:C6)", number_format, totals["current_total"])
    summary.write("A11", "Metric baseline total", label_format)
    summary.write_formula("B11", "=SUM(Metrics!D4:D6)", number_format, totals["baseline_total"])
    summary.write("A12", "Breached metric count", label_format)
    summary.write_formula("B12", '=COUNTIF(Metrics!F4:F6,"breached")', integer_format, totals["breached_count"])
    summary.write("A14", "Workbook note", label_format)
    summary.write(
        "B14",
        "This workbook is a stakeholder view over checked memo evidence, not a place for ad-hoc recalculation.",
    )
    summary.set_column("A:A", 26)
    summary.set_column("B:B", 95)

    metrics_sheet = workbook.add_worksheet("Metrics")
    metrics_sheet.write("A1", "Stakeholder metric summary", title_format)
    metric_rows: list[dict[str, Any]] = []
    for row in metrics:
        rendered = dict(row)
        for field in ("current", "baseline", "threshold"):
            try:
                rendered[field] = as_number(row.get(field, ""), field=field)
            except ValueError:
                rendered[field] = row.get(field, "")
        metric_rows.append(rendered)
    write_table(
        metrics_sheet,
        start_row=2,
        start_col=0,
        table_name="MetricsTable",
        columns=REQUIRED_METRIC_COLUMNS,
        rows=metric_rows,
        header_format=header_format,
    )
    metrics_sheet.set_column("A:A", 34)
    metrics_sheet.set_column("B:B", 36)
    metrics_sheet.set_column("C:E", 13, number_format)
    metrics_sheet.set_column("F:G", 18)
    metrics_sheet.conditional_format("F4:F6", {"type": "text", "criteria": "containing", "value": "watch", "format": warning_format})
    metrics_sheet.conditional_format("F4:F6", {"type": "text", "criteria": "containing", "value": "breached", "format": breach_format})

    evidence_sheet = workbook.add_worksheet("Evidence")
    evidence_sheet.write("A1", "Claim-evidence matrix excerpt", title_format)
    write_table(
        evidence_sheet,
        start_row=2,
        start_col=0,
        table_name="EvidenceTable",
        columns=REQUIRED_EVIDENCE_COLUMNS,
        rows=evidence,
        header_format=header_format,
    )
    evidence_sheet.set_column("A:E", 30)

    dictionary_sheet = workbook.add_worksheet("Data Dictionary")
    dictionary_sheet.write("A1", "Data dictionary", title_format)
    dictionary_columns = ["sheet", "column", "description", "source", "expected_type", "sensitive"]
    write_table(
        dictionary_sheet,
        start_row=2,
        start_col=0,
        table_name="DictionaryTable",
        columns=dictionary_columns,
        rows=dictionary_rows,
        header_format=header_format,
    )
    dictionary_sheet.set_column("A:B", 22)
    dictionary_sheet.set_column("C:C", 58)
    dictionary_sheet.set_column("D:F", 22)

    checks_sheet = workbook.add_worksheet("Checks")
    checks_sheet.write("A1", "Workbook audit checks", title_format)
    checks_rows = [
        {
            "check_id": item["id"],
            "valid": str(item["valid"]).lower(),
            "severity": item["severity"],
            "message": item["message"],
            "observed": json.dumps(item["observed"], ensure_ascii=False),
            "expected": json.dumps(item["expected"], ensure_ascii=False),
        }
        for item in initial_checks
    ]
    write_table(
        checks_sheet,
        start_row=2,
        start_col=0,
        table_name="ChecksTable",
        columns=["check_id", "valid", "severity", "message", "observed", "expected"],
        rows=checks_rows,
        header_format=header_format,
    )
    checks_sheet.set_column("A:A", 38)
    checks_sheet.set_column("B:C", 12)
    checks_sheet.set_column("D:D", 64)
    checks_sheet.set_column("E:F", 28)
    workbook.close()


def workbook_tables(workbook_path: str | Path) -> dict[str, str]:
    workbook = load_workbook(workbook_path, data_only=False)
    tables: dict[str, str] = {}
    for worksheet in workbook.worksheets:
        for table in worksheet.tables.values():
            tables[table.name] = table.ref
    return tables


def audit_workbook(
    *,
    workbook_path: str | Path,
    spec: dict[str, Any],
    metrics: list[dict[str, str]],
    evidence: list[dict[str, str]],
    initial_checks: list[dict[str, Any]],
) -> dict[str, Any]:
    workbook_file = Path(workbook_path)
    checks = list(initial_checks)
    if not workbook_file.is_file():
        checks.append(
            check(
                "workbook_file_exists",
                False,
                observed=str(workbook_file),
                expected="existing .xlsx file",
                message="Workbook file must be written before audit.",
            )
        )
        return build_audit(checks, workbook_id=str(spec.get("workbook_id", "")))

    formulas_wb = load_workbook(workbook_file, data_only=False)
    values_wb = load_workbook(workbook_file, data_only=True)
    sheet_names = formulas_wb.sheetnames
    checks.append(
        check(
            "required_sheets_present",
            sheet_names == REQUIRED_SHEETS,
            observed=sheet_names,
            expected=REQUIRED_SHEETS,
            message="Stakeholder workbook should keep a stable sheet order.",
        )
    )

    tables = workbook_tables(workbook_file)
    missing_tables = sorted(set(REQUIRED_TABLES.values()) - set(tables))
    checks.append(
        check(
            "required_excel_tables_present",
            not missing_tables,
            observed=missing_tables,
            expected=sorted(REQUIRED_TABLES.values()),
            message="Tables make stakeholder sheets filterable and auditable.",
        )
    )

    freeze_panes = {
        sheet: formulas_wb[sheet].freeze_panes
        for sheet in REQUIRED_TABLES
        if sheet in formulas_wb.sheetnames
    }
    checks.append(
        check(
            "table_sheets_are_frozen",
            all(value == "A4" for value in freeze_panes.values()),
            observed=freeze_panes,
            expected={sheet: "A4" for sheet in REQUIRED_TABLES},
            message="Header rows should stay visible while stakeholders scroll.",
        )
    )

    summary_formulas = {
        "B10": formulas_wb["Summary"]["B10"].value,
        "B11": formulas_wb["Summary"]["B11"].value,
        "B12": formulas_wb["Summary"]["B12"].value,
    }
    expected_formulas = {
        "B10": "=SUM(Metrics!C4:C6)",
        "B11": "=SUM(Metrics!D4:D6)",
        "B12": '=COUNTIF(Metrics!F4:F6,"breached")',
    }
    checks.append(
        check(
            "summary_formulas_are_present",
            summary_formulas == expected_formulas,
            observed=summary_formulas,
            expected=expected_formulas,
            message="Summary totals should be formulas, not pasted values.",
        )
    )

    totals = metric_totals(metrics)
    def cached_float(cell: str) -> float | None:
        value = values_wb["Summary"][cell].value
        if value is None:
            return None
        return round(float(value), 6)

    def cached_int(cell: str) -> int | None:
        value = values_wb["Summary"][cell].value
        if value is None:
            return None
        return int(value)

    cached_values = {
        "B10": cached_float("B10"),
        "B11": cached_float("B11"),
        "B12": cached_int("B12"),
    }
    expected_values = {
        "B10": totals["current_total"],
        "B11": totals["baseline_total"],
        "B12": totals["breached_count"],
    }
    checks.append(
        check(
            "summary_cached_totals_match_sources",
            cached_values == expected_values,
            observed=cached_values,
            expected=expected_values,
            message="Cached formula values must reconcile with source tables for reviewers without recalculation.",
        )
    )

    metrics_ws = values_wb["Metrics"]
    metric_ids = [
        metrics_ws.cell(row=row_number, column=1).value
        for row_number in range(4, 4 + len(metrics))
    ]
    checks.append(
        check(
            "metrics_sheet_preserves_source_order",
            metric_ids == [row["metric_id"] for row in metrics],
            observed=metric_ids,
            expected=[row["metric_id"] for row in metrics],
            message="Workbook rows should map back to source artifacts without guessing.",
        )
    )

    evidence_ws = values_wb["Evidence"]
    evidence_ids = [
        evidence_ws.cell(row=row_number, column=2).value
        for row_number in range(4, 4 + len(evidence))
    ]
    checks.append(
        check(
            "evidence_sheet_preserves_source_order",
            evidence_ids == [row["evidence_id"] for row in evidence],
            observed=evidence_ids,
            expected=[row["evidence_id"] for row in evidence],
            message="Evidence lineage should survive the workbook projection.",
        )
    )

    dictionary_ws = values_wb["Data Dictionary"]
    dictionary_pairs = {
        (dictionary_ws.cell(row=row_number, column=1).value, dictionary_ws.cell(row=row_number, column=2).value)
        for row_number in range(4, dictionary_ws.max_row + 1)
    }
    required_pairs = {
        ("Metrics", column) for column in REQUIRED_METRIC_COLUMNS
    } | {("Evidence", column) for column in REQUIRED_EVIDENCE_COLUMNS}
    missing_dictionary_pairs = sorted(
        f"{sheet}.{column}" for sheet, column in required_pairs - dictionary_pairs
    )
    checks.append(
        check(
            "workbook_dictionary_covers_visible_columns",
            not missing_dictionary_pairs,
            observed=missing_dictionary_pairs,
            expected=[],
            message="The workbook itself must include a complete data dictionary sheet.",
        )
    )

    return build_audit(checks, workbook_id=spec.get("workbook_id", ""))


def build_manifest(
    *,
    spec_path: Path,
    metrics_path: Path,
    evidence_path: Path,
    memo_audit_path: Path,
    output_paths: dict[str, Path],
) -> dict[str, Any]:
    return {
        "version": BUILDER_VERSION,
        "hash_algorithm": "sha256",
        "inputs": {
            "workbook_spec": {"path": str(spec_path), "sha256": sha256_file(spec_path)},
            "metric_summary": {"path": str(metrics_path), "sha256": sha256_file(metrics_path)},
            "claim_evidence_matrix": {
                "path": str(evidence_path),
                "sha256": sha256_file(evidence_path),
            },
            "memo_audit": {"path": str(memo_audit_path), "sha256": sha256_file(memo_audit_path)},
        },
        "outputs": {
            name: {"path": path.name, "sha256": sha256_file(path)}
            for name, path in sorted(output_paths.items())
        },
    }


def build_stakeholder_workbook(
    *,
    spec_path: str | Path,
    metrics_path: str | Path,
    evidence_path: str | Path,
    memo_audit_path: str | Path,
    output_dir: str | Path,
) -> WorkbookBuildResult:
    spec_file = Path(spec_path)
    metrics_file = Path(metrics_path)
    evidence_file = Path(evidence_path)
    memo_audit_file = Path(memo_audit_path)
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    spec = read_json(spec_file)
    metrics = read_csv(metrics_file)
    evidence = read_csv(evidence_file)
    memo_audit = read_json(memo_audit_file)
    input_checks = validate_inputs(spec, metrics, evidence, memo_audit)

    workbook_path = out / "stakeholder_workbook.xlsx"
    audit_path = out / "workbook_audit.json"
    dictionary_path = out / "data_dictionary.csv"
    manifest_path = out / "manifest.json"
    dictionary_rows = list(spec.get("data_dictionary", []))
    write_csv(
        dictionary_path,
        dictionary_rows,
        ["sheet", "column", "description", "source", "expected_type", "sensitive"],
    )
    render_workbook(workbook_path, spec, metrics, evidence, dictionary_rows, input_checks)
    audit = audit_workbook(
        workbook_path=workbook_path,
        spec=spec,
        metrics=metrics,
        evidence=evidence,
        initial_checks=input_checks,
    )
    write_json(audit_path, audit)
    manifest = build_manifest(
        spec_path=spec_file,
        metrics_path=metrics_file,
        evidence_path=evidence_file,
        memo_audit_path=memo_audit_file,
        output_paths={
            "stakeholder_workbook": workbook_path,
            "workbook_audit": audit_path,
            "data_dictionary": dictionary_path,
        },
    )
    write_json(manifest_path, manifest)
    return WorkbookBuildResult(
        output_dir=out,
        workbook_path=workbook_path,
        audit_path=audit_path,
        dictionary_path=dictionary_path,
        manifest_path=manifest_path,
        audit=audit,
    )


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build a stakeholder XLSX workbook package.")
    parser.add_argument("--spec", type=Path, help="Path to workbook_spec.json.")
    parser.add_argument("--metrics", type=Path, help="Path to metric_summary.csv.")
    parser.add_argument("--evidence", type=Path, help="Path to claim_evidence_matrix.csv.")
    parser.add_argument("--memo-audit", type=Path, help="Path to memo_audit.json from 17/01.")
    parser.add_argument("--output-dir", type=Path, required=True, help="Directory for workbook outputs.")
    parser.add_argument("--write-example", type=Path, help="Write sample inputs before building.")
    parser.add_argument(
        "--fail-on-invalid",
        action="store_true",
        help="Return non-zero when workbook audit is invalid.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_argument_parser()
    args = parser.parse_args(argv)
    if args.write_example:
        paths = write_sample_inputs(args.write_example)
        spec_path = paths["spec_path"]
        metrics_path = paths["metrics_path"]
        evidence_path = paths["evidence_path"]
        memo_audit_path = paths["memo_audit_path"]
    else:
        missing = [
            name
            for name, value in (
                ("--spec", args.spec),
                ("--metrics", args.metrics),
                ("--evidence", args.evidence),
                ("--memo-audit", args.memo_audit),
            )
            if value is None
        ]
        if missing:
            parser.error("missing required arguments without --write-example: " + ", ".join(missing))
        spec_path = args.spec
        metrics_path = args.metrics
        evidence_path = args.evidence
        memo_audit_path = args.memo_audit
    result = build_stakeholder_workbook(
        spec_path=spec_path,
        metrics_path=metrics_path,
        evidence_path=evidence_path,
        memo_audit_path=memo_audit_path,
        output_dir=args.output_dir,
    )
    report = {
        "valid": result.audit["valid"],
        "readiness_status": result.audit["readiness_status"],
        "workbook_path": str(result.workbook_path),
        "audit_path": str(result.audit_path),
        "dictionary_path": str(result.dictionary_path),
        "manifest_path": str(result.manifest_path),
        "blocking_errors": result.audit["summary"]["blocking_errors"],
    }
    print(json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True))
    if args.fail_on_invalid and not result.audit["valid"]:
        return 2
    return 0 if result.audit["valid"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
